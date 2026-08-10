"""
Servicio de exportación del inventario de equipos a Excel (openpyxl, modo
normal — `write_only` no soporta formato condicional ni `auto_filter`).

Genera un workbook con:
  - Hoja "Inventario": TODOS los resultados de aplicar los mismos filtros/scope
    que la lista de equipos (página + API JSON), con encabezado con estilo,
    panel congelado (`A2`), autofiltro y formato condicional tipo semáforo
    (garantía/verificación/estado) — nativo de Excel, no celda por celda.
  - Hoja "Resumen": bloques de conteo (estado, categoría, departamento,
    asignación, garantía, verificación) con un `BarChart` por bloque, más el
    total exportado y los filtros activos en texto legible.

SIN hoja de metadatos (decisión explícita del usuario). Por encima de
`_MAX_CONDITIONAL_ROWS` filas se DEGRADA (se omite el formato condicional,
que es lo que hace lento a Excel con muchas filas) en vez de rechazar la
exportación — se anota en la hoja de resumen.
"""
from __future__ import annotations

from datetime import date, datetime, timedelta
from io import BytesIO

from sqlalchemy.orm import Session, joinedload

# Por encima de este número de filas se omite el formato condicional (degrada,
# no rechaza — "todos los resultados de la búsqueda" es requisito explícito).
_MAX_CONDITIONAL_ROWS = 20_000

_HEADER_FILL_COLOR = "1F497D"

_RED_FILL = "FFC7CE"
_AMBER_FILL = "FFEB9C"
_GRAY_FILL = "D9D9D9"

# (encabezado, índice de columna 1-based) — usado para anchos + tipos de dato.
_HEADERS = [
    "N° Inventario", "Categoría", "Marca", "Modelo", "Serial Proveedor",
    "Serial ITCJ", "ID TecNM", "Departamento", "Asignado a", "Ubicación",
    "Estado", "Grupo", "Adquisición", "Vence garantía", "Días de garantía",
    "Último mant.", "Próximo mant.", "Última verificación",
    "Días sin verificar", "Registrado por", "Registrado el",
]
_COL_WIDTHS = [16, 18, 14, 16, 18, 18, 14, 20, 20, 18, 14, 16, 13, 13, 12, 13, 13, 16, 14, 20, 17]
# Columnas (1-based) que llevan formato de fecha dd/mm/yyyy.
_DATE_COLS = {13, 14, 16, 17, 18, 21}
# Letras de columna usadas por el formato condicional (fijas: el orden de
# _HEADERS arriba es la fuente de verdad).
_COL_STATUS = "K"
_COL_WARRANTY_DATE = "N"
_COL_VERIFIED_DATE = "R"
_COL_VERIFIED_DAYS = "S"

_WARRANTY_LABELS = {
    "valid": "Vigente", "expiring": "Por vencer (≤30 días)",
    "expired": "Vencida", "none": "Sin información",
}
_VERIFIED_LABELS = {
    "never": "Nunca verificado", "recent": "Reciente (<30 días)",
    "outdated": "Vencida (30-90 días)", "critical": "Crítica (>90 días)",
}
_MAINTENANCE_LABELS = {"due": "Vencido", "soon": "Próximo (≤30 días)"}
_SORT_LABELS = {
    "recent": "Más recientes", "oldest": "Más antiguos", "number": "N° Inventario",
    "warranty": "Garantía próxima", "verified": "Verificación pendiente",
}
_ASSIGNED_LABELS = {"yes": "Asignados", "no": "Globales"}


class InventoryExportService:
    """Exportación de la lista de equipos del inventario a Excel."""

    @staticmethod
    def export_items(db: Session, user: dict, params) -> tuple[BytesIO, str]:
        """Consulta (mismo scope/filtros que la lista) + arma el workbook.

        `params`: cualquier objeto con `.get(key)` — `request.query_params`
        (Starlette) o un `dict` plano (para llamarlo directo en tests, sin HTTP).
        Devuelve `(buffer_posicionado_en_0, nombre_de_archivo)`.
        """
        items, filters_display = InventoryExportService._query_items(db, user, params)
        wb = InventoryExportService._build_workbook(items, filters_display)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return buf, filename

    # ------------------------------------------------------------------
    # Query (mismo scope + mismos filtros que la lista)
    # ------------------------------------------------------------------

    @staticmethod
    def _query_items(db: Session, user: dict, params):
        from itcj2.apps.helpdesk.api.inventory.items import (
            _apply_item_common_filters,
            _apply_item_department_scope,
            _apply_item_sort,
        )
        from itcj2.apps.helpdesk.models.inventory_item import InventoryItem
        from itcj2.apps.helpdesk.utils.inventory_access import visible_department_ids

        user_id = int(user["sub"])

        def _s(key):
            v = params.get(key)
            if v is None:
                return None
            v = str(v).strip()
            return v or None

        search = _s("search")
        category_raw = _s("category")
        category_id = int(category_raw) if category_raw and category_raw.isdigit() else None
        status = _s("status")
        assigned = _s("assigned")
        dept_raw = _s("department")
        department_id = int(dept_raw) if dept_raw and dept_raw.isdigit() else None
        reg_start = _s("reg_start")
        reg_end = _s("reg_end")
        warranty = _s("warranty")
        verified = _s("verified")
        brand = _s("brand")
        maintenance = _s("maintenance")
        sort = _s("sort")

        # Mismo scope que GET /inventory/items: ?department= NUNCA sustituye el
        # scope visible, solo lo intersecta (fuera de scope ⇒ vacío).
        visible = visible_department_ids(db, user)

        query = (
            db.query(InventoryItem)
            .filter_by(is_active=True)
            .options(
                joinedload(InventoryItem.category),
                joinedload(InventoryItem.department),
                joinedload(InventoryItem.assigned_to_user),
                joinedload(InventoryItem.group),
                joinedload(InventoryItem.registered_by),
            )
        )
        query = _apply_item_department_scope(query, visible, department_id, user_id)
        query = _apply_item_common_filters(
            query,
            category_id=category_id, status=status, assigned=assigned, search=search,
            reg_start=reg_start, reg_end=reg_end, warranty=warranty, verified=verified,
            brand=brand, maintenance=maintenance,
        )
        query = _apply_item_sort(query, sort)

        items = query.all()

        filters_display = InventoryExportService._describe_filters(
            db,
            search=search, category_id=category_id, status=status, assigned=assigned,
            department_id=department_id, reg_start=reg_start, reg_end=reg_end,
            warranty=warranty, verified=verified, brand=brand, maintenance=maintenance,
            sort=sort,
        )
        return items, filters_display

    @staticmethod
    def _describe_filters(db: Session, **raw) -> list[str]:
        from itcj2.apps.helpdesk.models.inventory_category import InventoryCategory
        from itcj2.core.models.department import Department

        parts: list[str] = []
        if raw.get("search"):
            parts.append(f"Buscar: {raw['search']}")
        if raw.get("category_id"):
            cat = db.get(InventoryCategory, raw["category_id"])
            parts.append(f"Categoría: {cat.name if cat else raw['category_id']}")
        if raw.get("status"):
            parts.append(f"Estado: {raw['status'].upper()}")
        if raw.get("assigned"):
            parts.append(f"Asignación: {_ASSIGNED_LABELS.get(raw['assigned'].lower(), raw['assigned'])}")
        if raw.get("department_id"):
            dept = db.get(Department, raw["department_id"])
            parts.append(f"Departamento: {dept.name if dept else raw['department_id']}")
        if raw.get("reg_start"):
            parts.append(f"Registrado desde: {raw['reg_start']}")
        if raw.get("reg_end"):
            parts.append(f"Registrado hasta: {raw['reg_end']}")
        if raw.get("warranty"):
            parts.append(f"Garantía: {_WARRANTY_LABELS.get(raw['warranty'], raw['warranty'])}")
        if raw.get("verified"):
            parts.append(f"Verificación: {_VERIFIED_LABELS.get(raw['verified'], raw['verified'])}")
        if raw.get("brand"):
            parts.append(f"Marca: {raw['brand']}")
        if raw.get("maintenance"):
            parts.append(f"Mantenimiento: {_MAINTENANCE_LABELS.get(raw['maintenance'], raw['maintenance'])}")
        if raw.get("sort"):
            parts.append(f"Orden: {_SORT_LABELS.get(raw['sort'], raw['sort'])}")
        return parts

    # ------------------------------------------------------------------
    # Workbook
    # ------------------------------------------------------------------

    @staticmethod
    def _build_workbook(items: list, filters_display: list[str]):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"
        degraded = len(items) > _MAX_CONDITIONAL_ROWS
        InventoryExportService._write_items_sheet(ws, items, apply_conditional=not degraded)

        summary_ws = wb.create_sheet("Resumen")
        InventoryExportService._write_summary_sheet(summary_ws, items, filters_display, degraded=degraded)

        return wb

    @staticmethod
    def _write_items_sheet(ws, items: list, *, apply_conditional: bool) -> None:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", fgColor=_HEADER_FILL_COLOR)
        header_align = Alignment(horizontal="center", vertical="center")
        for col_idx, header in enumerate(_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        today = date.today()
        now = datetime.now()

        for row_idx, item in enumerate(items, start=2):
            warranty_days = (item.warranty_expiration - today).days if item.warranty_expiration else None
            verified_days = (now - item.last_verified_at).days if item.last_verified_at else None

            values = [
                item.inventory_number,
                item.category.name if item.category else None,
                item.brand,
                item.model,
                item.supplier_serial,
                item.itcj_serial,
                item.id_tecnm,
                item.department.name if item.department else None,
                item.assigned_to_user.full_name if item.assigned_to_user else "Global",
                item.location_detail,
                item.status,
                item.group.name if item.group else None,
                item.acquisition_date,
                item.warranty_expiration,
                warranty_days,
                item.last_maintenance_date,
                item.next_maintenance_date,
                item.last_verified_at,
                verified_days,
                item.registered_by.full_name if item.registered_by else None,
                item.registered_at,
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx in _DATE_COLS and value is not None:
                    cell.number_format = "dd/mm/yyyy"

        last_row = len(items) + 1
        last_col_letter = get_column_letter(len(_HEADERS))
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

        for idx, width in enumerate(_COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        if apply_conditional and items:
            InventoryExportService._apply_conditional_formatting(ws, len(items))

    @staticmethod
    def _apply_conditional_formatting(ws, n_items: int) -> None:
        """Semáforos NATIVOS de Excel (Formato condicional real, no relleno
        celda-por-celda en Python) — así el usuario puede editarlos/ajustarlos
        directo en Excel y no penaliza al escribir archivos grandes."""
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import PatternFill

        red = PatternFill(start_color=_RED_FILL, end_color=_RED_FILL, fill_type="solid")
        amber = PatternFill(start_color=_AMBER_FILL, end_color=_AMBER_FILL, fill_type="solid")
        gray = PatternFill(start_color=_GRAY_FILL, end_color=_GRAY_FILL, fill_type="solid")

        last_row = n_items + 1
        c = _COL_WARRANTY_DATE
        rng = f"{c}2:{c}{last_row}"
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND({c}2<>"",{c}2<TODAY())'], fill=red))
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'AND({c}2<>"",{c}2>=TODAY(),{c}2<=TODAY()+30)'], fill=amber,
        ))

        days_col = _COL_VERIFIED_DAYS
        date_col = _COL_VERIFIED_DATE
        rng = f"{days_col}2:{days_col}{last_row}"
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{date_col}2=""'], fill=red))
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'AND({date_col}2<>"",{days_col}2>90)'], fill=red,
        ))
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'AND({date_col}2<>"",{days_col}2>30,{days_col}2<=90)'], fill=amber,
        ))

        c = _COL_STATUS
        rng = f"{c}2:{c}{last_row}"
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'OR({c}2="DAMAGED",{c}2="LOST")'], fill=red))
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{c}2="MAINTENANCE"'], fill=amber))
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{c}2="RETIRED"'], fill=gray))

    # ------------------------------------------------------------------
    # Resumen
    # ------------------------------------------------------------------

    @staticmethod
    def _write_summary_sheet(ws, items: list, filters_display: list[str], *, degraded: bool) -> None:
        from openpyxl.chart import BarChart, Reference
        from openpyxl.styles import Font

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 14

        row = 1
        ws.cell(row=row, column=1, value="Resumen de exportación de inventario").font = Font(bold=True, size=13)
        row += 1
        ws.cell(row=row, column=1, value=f"Total exportado: {len(items)}").font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value=(
            "Filtros activos: " + "; ".join(filters_display) if filters_display
            else "Filtros activos: ninguno (todos los equipos visibles)"
        ))
        row += 1
        if degraded:
            cell = ws.cell(row=row, column=1, value=(
                f"Más de {_MAX_CONDITIONAL_ROWS:,} filas: el formato condicional (semáforos) "
                "se omitió en la hoja Inventario para mantener el archivo ligero."
            ))
            cell.font = Font(color="9C5700", italic=True)
            row += 1
        row += 1  # línea en blanco

        counts = InventoryExportService._compute_summary_counts(items)

        min_gap = 16  # filas mínimas entre el inicio de un bloque y el siguiente (deja hueco al chart)
        for title, block_counts in counts:
            start_row = row
            ws.cell(row=row, column=1, value=title).font = Font(bold=True)
            row += 1
            header_row = row
            ws.cell(row=row, column=1, value="Concepto")
            ws.cell(row=row, column=2, value="Cantidad")
            row += 1
            data_start = row
            for label, count in block_counts.items():
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=count)
                row += 1
            data_end = row - 1

            if data_end >= data_start:
                chart = BarChart()
                chart.type = "col"
                chart.title = title
                chart.height = 6
                chart.width = 12
                chart.legend = None
                data_ref = Reference(ws, min_col=2, min_row=header_row, max_row=data_end)
                cats_ref = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                ws.add_chart(chart, f"D{start_row}")

            row = max(row + 1, start_row + min_gap)

    @staticmethod
    def _compute_summary_counts(items: list) -> list[tuple[str, dict]]:
        from itcj2.apps.helpdesk.api.inventory.verification import _verification_status

        today = date.today()

        status_counts: dict[str, int] = {}
        category_counts: dict[str, int] = {}
        dept_counts: dict[str, int] = {}
        assign_counts = {"Asignados": 0, "Globales": 0}
        warranty_counts = {"Vigente": 0, "Por vencer (≤30 días)": 0, "Vencida": 0, "Sin información": 0}
        verif_counts = {"Reciente (<30 días)": 0, "Vencida (30-90 días)": 0, "Crítica (>90 días)": 0, "Nunca verificado": 0}

        for item in items:
            status_counts[item.status] = status_counts.get(item.status, 0) + 1

            cat_name = item.category.name if item.category else "Sin categoría"
            category_counts[cat_name] = category_counts.get(cat_name, 0) + 1

            dept_name = item.department.name if item.department else "Sin departamento"
            dept_counts[dept_name] = dept_counts.get(dept_name, 0) + 1

            if item.assigned_to_user_id:
                assign_counts["Asignados"] += 1
            else:
                assign_counts["Globales"] += 1

            if not item.warranty_expiration:
                warranty_counts["Sin información"] += 1
            elif item.warranty_expiration < today:
                warranty_counts["Vencida"] += 1
            elif item.warranty_expiration <= today + timedelta(days=30):
                warranty_counts["Por vencer (≤30 días)"] += 1
            else:
                warranty_counts["Vigente"] += 1

            vs = _verification_status(item.last_verified_at)
            if vs == "never":
                verif_counts["Nunca verificado"] += 1
            elif vs == "recent":
                verif_counts["Reciente (<30 días)"] += 1
            elif vs == "outdated":
                verif_counts["Vencida (30-90 días)"] += 1
            else:
                verif_counts["Crítica (>90 días)"] += 1

        return [
            ("Por estado", status_counts),
            ("Por categoría", category_counts),
            ("Por departamento", dept_counts),
            ("Asignación", assign_counts),
            ("Garantía", warranty_counts),
            ("Verificación", verif_counts),
        ]
