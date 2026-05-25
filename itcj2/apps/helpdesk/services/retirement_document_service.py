"""
Servicio de generación del oficio de baja editando la plantilla oficial
`OFICIO PARA BAJAS-.xlsx`. Sólo se edita el archivo provisto — no se
genera PDF para preservar el formato institucional.

Auto-rellena:
  - Folio en N6 (ANEXO AL OFICIO)
  - Fecha en M8 (FECHA)
  - Ubicación (nombre del depto) en D10
  - U.R. = 513 en H10 (fijo)
  - Causa de baja en N10
  - Responsable (jefe del depto del item) en C12
  - HOJA 1 DE 1 en M12/O12
  - Filas de items desde fila 18 (hasta 27 por hoja física)
  - Bloque de firmas en filas 47-54 (Director / Subdirector / Jefe Rec Mat /
    Recibe Bodega / Registro y Control) con nombre + fecha cuando ya hay firma
"""
import io
import logging
import os
from datetime import datetime
from typing import Optional

from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)

# ── Constantes institucionales ─────────────────────────────────────────────────

_INSTITUCION = "INSTITUTO TECNOLOGICO DE CD JUAREZ"
_CLAVE_CT    = "08DIT0014X"
_NUM_UR      = "513"

_SIGNER_POSITIONS = {
    1: ("head_mat_services",          "Jefe del Depto. de Recursos Materiales y Servicios"),
    2: ("subdirector_admin_services", "Subdirector de Servicios Administrativos"),
    3: ("director",                   "Director"),
}

_EXCEL_TEMPLATE = os.path.join(
    os.path.dirname(__file__), "..", "static", "data", "OFICIO PARA BAJAS-.xlsx",
)

_ITEMS_START_ROW   = 18
_ITEMS_END_ROW     = 44   # antes del TOTAL en B45 — máximo 27 items por hoja


# ── Helpers ────────────────────────────────────────────────────────────────────

def _resolve_signer(db: Session, code: str) -> Optional[dict]:
    """Devuelve {id, full_name, title} del primer usuario activo con ese position."""
    from itcj2.core.models.position import Position, UserPosition
    from itcj2.core.models.user import User

    row = (
        db.query(User, Position.title)
        .join(UserPosition, UserPosition.user_id == User.id)
        .join(Position, Position.id == UserPosition.position_id)
        .filter(
            Position.code == code,
            UserPosition.is_active.is_(True),
            User.is_active.is_(True),
        )
        .first()
    )
    if not row:
        return None
    user, title = row
    return {"id": user.id, "full_name": user.full_name, "title": title}


def _resolve_dept_head(db: Session, department_id: int) -> Optional[dict]:
    """Jefe del departamento (puesto que empieza por 'head_' del depto)."""
    from itcj2.core.models.position import Position, UserPosition
    from itcj2.core.models.user import User

    row = (
        db.query(User, Position.title)
        .join(UserPosition, UserPosition.user_id == User.id)
        .join(Position, Position.id == UserPosition.position_id)
        .filter(
            Position.department_id == department_id,
            Position.code.like('head_%'),
            UserPosition.is_active.is_(True),
            User.is_active.is_(True),
        )
        .order_by(Position.id.asc())
        .first()
    )
    if not row:
        return None
    user, title = row
    return {"id": user.id, "full_name": user.full_name, "title": title}


def _request_context(db: Session, request) -> dict:
    items = list(request.items.all())

    # Departamento principal (del primer item)
    dept_id = None
    dept_name = "—"
    for ri in items:
        if ri.inventory_item and ri.inventory_item.department:
            dept_id = ri.inventory_item.department.id
            dept_name = ri.inventory_item.department.name
            break

    responsable = _resolve_dept_head(db, dept_id) if dept_id else None

    signers = {}
    for step, (code, fallback_title) in _SIGNER_POSITIONS.items():
        info = _resolve_signer(db, code)
        if info is None:
            info = {"id": None, "full_name": None, "title": fallback_title}
        signers[step] = info

    sigs_by_step = {sig.step: sig for sig in (request.signatures or [])}

    return {
        "folio":              request.folio,
        "fecha_oficio":       request.created_at,
        "location_text":      dept_name,
        "causa":              request.reason or "—",
        "responsable":        responsable,
        "signers":            signers,
        "signatures_by_step": sigs_by_step,
        "items":              items,
    }


class RetirementDocumentService:
    """Genera oficio de baja editando la plantilla oficial Excel."""

    # ── Excel template fill ────────────────────────────────────────────────────

    @staticmethod
    def fill_excel_template(request, db: Optional[Session] = None) -> bytes:
        """Llena la plantilla oficial OFICIO PARA BAJAS-.xlsx y devuelve bytes."""
        try:
            import openpyxl
        except ImportError as exc:
            logger.error(f"openpyxl no disponible: {exc}")
            raise RuntimeError("openpyxl no está instalado") from exc

        if db is None:
            from sqlalchemy.orm import object_session
            db = object_session(request)

        ctx = _request_context(db, request) if db else None

        template_path = os.path.normpath(_EXCEL_TEMPLATE)
        if not os.path.exists(template_path):
            raise RuntimeError(f"Plantilla no encontrada en {template_path}")

        wb = openpyxl.load_workbook(template_path, keep_vba=False)
        ws = wb.active

        if ctx:
            RetirementDocumentService._fill_header(ws, ctx)
            RetirementDocumentService._fill_items(ws, ctx)
            RetirementDocumentService._fill_signers(ws, ctx)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── PDF — DESHABILITADO ────────────────────────────────────────────────────

    @staticmethod
    def generate_pdf(request, db=None) -> bytes:  # noqa: ARG004
        """PDF deshabilitado por instrucción del usuario; usar Excel."""
        raise RuntimeError(
            "Generación de PDF deshabilitada. Use formato Excel (xlsx) "
            "para mantener el formato oficial."
        )

    # ── Helpers internos de relleno ────────────────────────────────────────────

    @staticmethod
    def _fill_header(ws, ctx: dict):
        """Llena celdas conocidas del encabezado de la plantilla + restaura caja
        de la columna derecha (ANEXO / FECHA / CAUSA / HOJA) que openpyxl pierde
        si la plantilla original usa formas/shapes dibujadas en lugar de bordes
        de celda.
        """
        from openpyxl.styles import Border, Side, Alignment, Font

        fecha_str   = ctx["fecha_oficio"].strftime("%d/%m/%Y") if ctx["fecha_oficio"] else ""
        responsable = ctx["responsable"]["full_name"] if ctx["responsable"] else ""
        causa_text  = ctx["causa"] or "OBSOLETO"  # default si no se proporciona

        ws["N6"]  = ctx["folio"] or ""
        ws["M8"]  = fecha_str
        ws["D10"] = ctx["location_text"]
        ws["H10"] = _NUM_UR
        ws["N10"] = causa_text
        ws["C12"] = responsable
        ws["M12"] = "1"
        ws["O12"] = "1"

        # ── Restaurar caja del lado derecho (ANEXO / FECHA / HOJA) ─────────────
        # Replica el rectángulo redondeado del template original con bordes
        # de celda. Cubre L6:P12 (la sección a la derecha del encabezado).
        thin = Side(style="thin", color="000000")
        outer = Border(top=thin, bottom=thin, left=thin, right=thin)

        # Recorrer perímetro de L6:P12 dibujando borde grueso por celda
        for r in range(6, 13):
            for c in range(12, 17):  # L=12, P=16
                cell = ws.cell(row=r, column=c)
                # bordes individuales según posición
                left   = thin if c == 12 else cell.border.left
                right  = thin if c == 16 else cell.border.right
                top    = thin if r == 6  else cell.border.top
                bottom = thin if r == 12 else cell.border.bottom
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

        # ── Estilo del valor de causa para que se lea bien ─────────────────────
        c_causa = ws["N10"]
        c_causa.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c_causa.font = Font(size=9, bold=True)

    @staticmethod
    def _fill_items(ws, ctx: dict):
        """Coloca un item por fila comenzando en la fila 18.
        Texto centrado en todas las celdas; flags DESALOJO/BODEGA/AFECTACIÓN
        por defecto en SI (a menos que el usuario marque explícitamente NO).
        """
        from openpyxl.styles import Alignment, Font

        center  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        font_8  = Font(size=9)
        font_8b = Font(size=9, bold=True)

        max_items = _ITEMS_END_ROW - _ITEMS_START_ROW + 1
        for idx, ri in enumerate(ctx["items"][:max_items]):
            r = _ITEMS_START_ROW + idx
            inv_item = ri.inventory_item
            inv_number = inv_item.inventory_number if inv_item else ""
            descripcion = (
                f"MARCA {inv_item.brand or ''}, MODELO {inv_item.model or ''}".strip()
                if inv_item else ""
            )

            cells_to_set = [
                (1,  idx + 1,                                          font_8b),  # A: PROG
                (2,  inv_number,                                       font_8b),  # B: INVENTARIO
                (3,  1,                                                font_8),   # C: CANTIDAD
                (4,  descripcion,                                      font_8),   # D: DESCRIPCIÓN
                (5,  float(ri.valor_unitario) if ri.valor_unitario is not None else "", font_8),  # E: VALOR
                (6,  ri.item_notes or "",                              font_8),   # F: DIAGNÓSTICO
                # G/H DESALOJO   — default SI (sólo cae NO si flag está explícitamente False y user marcó así)
                (7,  "SI",                                             font_8b),  # G: SI
                (8,  "",                                               font_8),   # H: NO
                (9,  "SI",                                             font_8b),  # I: BODEGA SI
                (10, "",                                               font_8),   # J: NO
                (11, "SI",                                             font_8b),  # K: AFECT SI
                (12, "",                                               font_8),   # L: NO
                (16, ri.item_notes or "",                              font_8),   # P: OBSERVACIONES
            ]
            for col, val, ft in cells_to_set:
                c = ws.cell(row=r, column=col)
                c.value = val
                c.alignment = center
                c.font = ft

    @staticmethod
    def _fill_signers(ws, ctx: dict):
        """
        Escribe el bloque de firmas en filas 47-54.
        Layout: 5 columnas (DIRECTOR / SUBDIRECTOR / JEFE REC MAT / RECIBE BODEGA / REGISTRO),
        cada bloque ocupa 3 columnas del worksheet:
          DIRECTOR              → cols A-C   (1..3)
          SUBDIRECTOR           → cols D-F   (4..6)
          JEFE REC MAT          → cols G-I   (7..9)
          RECIBE BODEGA         → cols J-L   (10..12)
          REGISTRO Y CONTROL    → cols M-P   (13..16)
        """
        from openpyxl.styles import Font, Alignment, Border, Side

        signers = ctx["signers"]
        sigs_by_step = ctx["signatures_by_step"]

        col_groups = [
            (1,  3,  "DIRECTOR.",                                signers.get(3), sigs_by_step.get(3)),
            (4,  6,  "SUBDIRECTOR DE SERVICIOS ADMINISTRATIVOS", signers.get(2), sigs_by_step.get(2)),
            (7,  9,  "JEFE DE RECURSOS MATERIALES Y SERVICIOS",  signers.get(1), sigs_by_step.get(1)),
            (10, 12, "RECIBE BODEGA",                            None, None),
            (13, 16, "REGISTRO Y CONTROL",                       None, None),
        ]

        bold_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        regular_center = Alignment(horizontal="center", vertical="center")
        bold = Font(bold=True, size=8)
        regular = Font(size=8)
        approved = Font(bold=True, size=8, color="0a7d0a")
        rejected = Font(bold=True, size=8, color="C00000")

        thin = Side(style="thin", color="000000")
        # Caja completa rodeando bloque de firmas filas 46-53
        SIG_TOP_ROW    = 46
        SIG_BOTTOM_ROW = 53

        # Dibujar bordes externos y verticales separadores entre cargos
        for r in range(SIG_TOP_ROW, SIG_BOTTOM_ROW + 1):
            for (c_start, c_end, _, _, _) in col_groups:
                for c in range(c_start, c_end + 1):
                    cell = ws.cell(row=r, column=c)
                    left   = thin if c == c_start else cell.border.left
                    right  = thin if c == c_end   else cell.border.right
                    top    = thin if r == SIG_TOP_ROW    else cell.border.top
                    bottom = thin if r == SIG_BOTTOM_ROW else cell.border.bottom
                    cell.border = Border(left=left, right=right, top=top, bottom=bottom)

        # Encabezado (rótulo del cargo) — fila 47
        header_row = 47
        for (c_start, c_end, label, _, _) in col_groups:
            try:
                ws.unmerge_cells(start_row=header_row, start_column=c_start,
                                 end_row=header_row, end_column=c_end)
            except Exception:
                pass
            ws.merge_cells(start_row=header_row, start_column=c_start,
                           end_row=header_row, end_column=c_end)
            c = ws.cell(row=header_row, column=c_start)
            c.value = label
            c.font = bold
            c.alignment = bold_center
        ws.row_dimensions[header_row].height = 24

        # FECHA — fila 48
        fecha_row = 48
        for (c_start, c_end, _, _, sig_obj) in col_groups:
            date_val = ""
            if sig_obj and sig_obj.signed_at:
                date_val = f"FECHA: {sig_obj.signed_at.strftime('%d/%m/%Y')}"
            else:
                date_val = "FECHA: ______________"
            try:
                ws.unmerge_cells(start_row=fecha_row, start_column=c_start,
                                 end_row=fecha_row, end_column=c_end)
            except Exception:
                pass
            ws.merge_cells(start_row=fecha_row, start_column=c_start,
                           end_row=fecha_row, end_column=c_end)
            c = ws.cell(row=fecha_row, column=c_start)
            c.value = date_val
            c.font = regular
            c.alignment = regular_center

        # NOMBRE — fila 49 (2 filas de alto)
        nombre_row = 49
        ws.row_dimensions[nombre_row].height = 28
        for (c_start, c_end, _, signer, _) in col_groups:
            name = (signer or {}).get("full_name") if signer else None
            nom_val = f"NOMBRE: {name}" if name else "NOMBRE: ______________"
            try:
                ws.unmerge_cells(start_row=nombre_row, start_column=c_start,
                                 end_row=nombre_row, end_column=c_end)
            except Exception:
                pass
            ws.merge_cells(start_row=nombre_row, start_column=c_start,
                           end_row=nombre_row, end_column=c_end)
            c = ws.cell(row=nombre_row, column=c_start)
            c.value = nom_val
            c.font = Font(bold=True, size=8)
            c.alignment = bold_center

        # FIRMA: ___ — fila 51
        firma_row = 51
        ws.row_dimensions[firma_row].height = 26
        for (c_start, c_end, _, _, _) in col_groups:
            try:
                ws.unmerge_cells(start_row=firma_row, start_column=c_start,
                                 end_row=firma_row, end_column=c_end)
            except Exception:
                pass
            ws.merge_cells(start_row=firma_row, start_column=c_start,
                           end_row=firma_row, end_column=c_end)
            c = ws.cell(row=firma_row, column=c_start)
            c.value = "FIRMA: ________________"
            c.font = regular
            c.alignment = regular_center

        # Estado autorización — fila 52
        estado_row = 52
        for (c_start, c_end, _, signer, sig_obj) in col_groups:
            if signer is None:
                state_val = ""
                state_font = regular
            elif sig_obj is None or sig_obj.action is None:
                state_val = "(Pendiente)"
                state_font = Font(italic=True, size=7, color="808080")
            elif sig_obj.action == "APPROVED":
                state_val = "✓ AUTORIZADO"
                state_font = approved
            elif sig_obj.action == "REJECTED":
                state_val = "✗ RECHAZADO"
                state_font = rejected
            else:
                state_val = ""
                state_font = regular
            try:
                ws.unmerge_cells(start_row=estado_row, start_column=c_start,
                                 end_row=estado_row, end_column=c_end)
            except Exception:
                pass
            ws.merge_cells(start_row=estado_row, start_column=c_start,
                           end_row=estado_row, end_column=c_end)
            c = ws.cell(row=estado_row, column=c_start)
            c.value = state_val
            c.font = state_font
            c.alignment = regular_center
