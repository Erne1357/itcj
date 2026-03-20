"""
Tareas Celery del módulo Helpdesk.

Tareas disponibles:
    cleanup_attachments     — elimina adjuntos expirados y marca los de tickets cerrados
    convert_document        — genera solicitud/orden_trabajo como PDF en background
    export_inventory_report — exporta inventario a CSV o XLSX en background
"""
import json
import logging
import os

from itcj2.celery_app import celery_app
from itcj2.tasks.base import LoggedTask

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Metadata de registro (se usa en CLI sync-tasks para poblar TaskDefinition)
# ---------------------------------------------------------------------------

TASK_DEFINITIONS = [
    {
        "task_name": "itcj2.tasks.helpdesk_tasks.cleanup_attachments",
        "display_name": "Limpieza de Adjuntos Expirados",
        "description": (
            "Marca adjuntos de tickets resueltos/cerrados para auto-eliminación (7 días) "
            "y elimina los que ya pasaron su fecha de expiración del disco y la base de datos."
        ),
        "app_name": "helpdesk",
        "category": "maintenance",
        "default_args": {"dry_run": False},
    },
    {
        "task_name": "itcj2.tasks.helpdesk_tasks.convert_document",
        "display_name": "Generar Documento PDF",
        "description": (
            "Genera la Solicitud de Mantenimiento o la Orden de Trabajo de un ticket "
            "como PDF en segundo plano usando LibreOffice. Notifica al usuario cuando "
            "el archivo está listo para descargar."
        ),
        "app_name": "helpdesk",
        "category": "document",
        "default_args": {"ticket_id": 0, "doc_type": "solicitud", "notify_user_id": 0},
    },
    {
        "task_name": "itcj2.tasks.helpdesk_tasks.export_inventory_report",
        "display_name": "Exportar Reporte de Inventario",
        "description": (
            "Genera un reporte de equipos de inventario en formato CSV o XLSX, "
            "lo guarda en disco y notifica al usuario solicitante con el enlace "
            "de descarga."
        ),
        "app_name": "helpdesk",
        "category": "report",
        "default_args": {"filters": {}, "format": "xlsx", "requested_by_user_id": 0},
    },
]


# ---------------------------------------------------------------------------
# cleanup_attachments
# ---------------------------------------------------------------------------

@celery_app.task(
    bind=True,
    base=LoggedTask,
    name="itcj2.tasks.helpdesk_tasks.cleanup_attachments",
    max_retries=2,
    default_retry_delay=60,
    soft_time_limit=120,
    time_limit=150,
)
def cleanup_attachments(self, task_run_id: int | None = None, dry_run: bool = False):
    """
    Tarea de mantenimiento: limpia adjuntos del helpdesk.

    Paso 1 — Marcar para borrado: recorre tickets en status CLOSED sin fecha
              de auto-delete y asigna auto_delete_at = ticket.updated_at + 7 días
              (updated_at es la fecha de cierre, ya que un ticket cerrado no cambia más).
    Paso 2 — Eliminar expirados: borra archivos del disco y registros de DB
              donde auto_delete_at <= ahora Y el ticket lleva >= 7 días cerrado.

    Args:
        task_run_id: ID del TaskRun creado por la API antes de encolar esta tarea.
                     Puede ser None si el scheduler no pudo inyectarlo (modo degradado).
        dry_run: Si True, solo cuenta lo que haría sin modificar nada.

    Returns:
        dict con claves: marked_for_delete, deleted_files, freed_bytes, errors, dry_run,
                         by_ticket (desglose por número de ticket)
    """
    from itcj2.database import SessionLocal
    from itcj2.apps.helpdesk.services.attachment_cleanup import (
        set_auto_delete_on_closed_tickets,
    )

    logger.info(f"[cleanup_attachments] Iniciando (dry_run={dry_run}, task_run_id={task_run_id})")

    errors = []
    marked = 0
    deleted = 0
    freed_bytes = 0
    by_ticket: dict = {}

    try:
        with SessionLocal() as db:
            # Paso 1: marcar tickets resueltos
            self.update_progress(task_run_id, current=0, total=2, message="Marcando adjuntos de tickets cerrados...")
            if not dry_run:
                marked = set_auto_delete_on_closed_tickets(db)
            else:
                # En dry_run solo contar, sin commit
                from itcj2.apps.helpdesk.models.attachment import Attachment
                from itcj2.apps.helpdesk.models.ticket import Ticket
                closed_tickets = db.query(Ticket).filter(
                    Ticket.status == "CLOSED",
                ).all()
                for ticket in closed_tickets:
                    for att in ticket.attachments:
                        if att.auto_delete_at is None:
                            marked += 1

            # Paso 2: eliminar expirados
            self.update_progress(task_run_id, current=1, total=2, message="Eliminando adjuntos expirados...")
            if not dry_run:
                deleted, freed_bytes, by_ticket = _cleanup_with_metrics(db, errors)
            else:
                from itcj2.apps.helpdesk.models.attachment import Attachment
                from itcj2.apps.helpdesk.models.ticket import Ticket
                from itcj2.apps.helpdesk.services.attachment_cleanup import AUTO_DELETE_DAYS
                from datetime import datetime, timedelta
                _now = datetime.utcnow()
                _cutoff = _now - timedelta(days=AUTO_DELETE_DAYS)
                expired = (
                    db.query(Attachment)
                    .join(Ticket, Ticket.id == Attachment.ticket_id)
                    .filter(
                        Attachment.auto_delete_at.isnot(None),
                        Attachment.auto_delete_at <= _now,
                        Ticket.status == "CLOSED",
                        Ticket.updated_at <= _cutoff,
                    )
                    .all()
                )
                deleted = len(expired)
                import os
                freed_bytes = sum(
                    os.path.getsize(a.filepath)
                    for a in expired
                    if a.filepath and os.path.exists(a.filepath)
                )
                by_ticket = _build_ticket_breakdown(expired)

    except Exception as exc:
        logger.exception(f"[cleanup_attachments] Error inesperado: {exc}")
        raise self.retry(exc=exc)

    result = {
        "marked_for_delete": marked,
        "deleted_files": deleted,
        "freed_bytes": freed_bytes,
        "freed_mb": round(freed_bytes / (1024 * 1024), 2),
        "errors": errors,
        "dry_run": dry_run,
        "by_ticket": by_ticket,
    }
    _log_cleanup_result(result)
    return result


# ---------------------------------------------------------------------------
# Helpers compartidos
# ---------------------------------------------------------------------------

def _get_exports_dir() -> str:
    """Devuelve (y crea si no existe) el directorio de exports del helpdesk."""
    from itcj2.config import get_settings
    exports_dir = os.path.join(
        get_settings().INSTANCE_PATH, "apps", "helpdesk", "exports"
    )
    os.makedirs(exports_dir, exist_ok=True)
    return exports_dir


def _push_user_notification(user_id: int, title: str, body: str, link: str | None) -> None:
    """Crea un registro Notification en DB y publica en Redis para Socket.IO."""
    try:
        from itcj2.database import SessionLocal
        from itcj2.core.models.notification import Notification
        import redis
        from itcj2.config import get_settings

        with SessionLocal() as db:
            notif = Notification(
                user_id=user_id,
                app_name="helpdesk",
                type="SYSTEM",
                title=title,
                body=body,
                data={"url": link} if link else {},
            )
            db.add(notif)
            db.commit()
            db.refresh(notif)
            notif_dict = notif.to_dict()

        r = redis.from_url(get_settings().REDIS_URL)
        r.publish("task_events", json.dumps({
            "type": "user_notification",
            "user_id": user_id,
            "notification": notif_dict,
        }))
    except Exception as e:
        logger.error(f"[helpdesk_tasks] Error creando notificación para user {user_id}: {e}")


def _cleanup_with_metrics(db, errors: list) -> tuple:
    """Elimina adjuntos expirados y devuelve (deleted_count, freed_bytes, by_ticket).

    Doble condición de seguridad:
      1. auto_delete_at está fijado y ya venció.
      2. El ticket está CLOSED y ticket.updated_at tiene >= 7 días (guarda
         contra fechas mal calculadas que pudieran adelantar el borrado).
    """
    import os
    from datetime import datetime, timedelta
    from itcj2.apps.helpdesk.models.attachment import Attachment
    from itcj2.apps.helpdesk.models.ticket import Ticket
    from itcj2.apps.helpdesk.services.attachment_cleanup import AUTO_DELETE_DAYS

    now = datetime.utcnow()
    cutoff = now - timedelta(days=AUTO_DELETE_DAYS)

    expired = (
        db.query(Attachment)
        .join(Ticket, Ticket.id == Attachment.ticket_id)
        .filter(
            Attachment.auto_delete_at.isnot(None),
            Attachment.auto_delete_at <= now,
            Ticket.status == "CLOSED",
            Ticket.updated_at <= cutoff,
        )
        .all()
    )

    deleted = 0
    freed_bytes = 0
    from datetime import timezone
    deletion_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    for attachment in expired:
        try:
            _append_deletion_audit_note(db, attachment, deletion_date)
            if attachment.filepath and os.path.exists(attachment.filepath):
                freed_bytes += os.path.getsize(attachment.filepath)
                os.remove(attachment.filepath)
            db.delete(attachment)
            deleted += 1
        except Exception as e:
            errors.append({"attachment_id": attachment.id, "error": str(e)})
            logger.error(f"Error eliminando attachment {attachment.id}: {e}")

    if deleted > 0:
        db.commit()

    by_ticket = _build_ticket_breakdown(expired)
    return deleted, freed_bytes, by_ticket


def _build_ticket_breakdown(attachments: list) -> dict:
    """Agrupa los adjuntos eliminados por número de ticket.

    Returns:
        dict[ticket_number, {ticket_image, resolution, comment, total, freed_bytes}]
    """
    import os
    breakdown: dict = {}

    for att in attachments:
        ticket = att.ticket
        key = ticket.ticket_number if ticket else f"ticket_{att.ticket_id}"

        if key not in breakdown:
            breakdown[key] = {
                "ticket_image": 0,
                "resolution": 0,
                "comment": 0,
                "total": 0,
                "freed_bytes": 0,
            }

        att_type = att.attachment_type
        if att_type == "ticket":
            breakdown[key]["ticket_image"] += 1
        elif att_type == "resolution":
            breakdown[key]["resolution"] += 1
        elif att_type == "comment":
            breakdown[key]["comment"] += 1

        breakdown[key]["total"] += 1

        if att.filepath and os.path.exists(att.filepath):
            try:
                breakdown[key]["freed_bytes"] += os.path.getsize(att.filepath)
            except OSError:
                pass

    return breakdown


def _log_cleanup_result(result: dict) -> None:
    """Emite un log detallado del resultado de la limpieza de adjuntos."""
    prefix = "[cleanup_attachments][DRY RUN]" if result.get("dry_run") else "[cleanup_attachments]"

    logger.info(
        "%s Completado — marcados: %d | eliminados: %d | liberados: %.2f MB | errores: %d",
        prefix,
        result["marked_for_delete"],
        result["deleted_files"],
        result["freed_mb"],
        len(result["errors"]),
    )

    by_ticket: dict = result.get("by_ticket") or {}
    if not by_ticket:
        logger.info("%s Sin adjuntos eliminados en esta ejecución.", prefix)
        return

    logger.info("%s Desglose por ticket (%d tickets afectados):", prefix, len(by_ticket))
    for ticket_number, data in sorted(by_ticket.items()):
        freed_mb = round(data["freed_bytes"] / (1024 * 1024), 2)
        parts = []
        if data["ticket_image"]:
            parts.append(f"imagen_ticket={data['ticket_image']}")
        if data["resolution"]:
            parts.append(f"resolución={data['resolution']}")
        if data["comment"]:
            parts.append(f"comentario={data['comment']}")
        logger.info(
            "%s   %s → total=%d (%s) | %.2f MB",
            prefix,
            ticket_number,
            data["total"],
            ", ".join(parts) if parts else "sin tipo clasificado",
            freed_mb,
        )


def _append_deletion_audit_note(db, attachment, deletion_date: str) -> None:
    """
    Antes de borrar un adjunto, deja rastro en el texto de su registro padre:
      - 'ticket'     → ticket.description       (solo indica que había una imagen)
      - 'resolution' → ticket.resolution_notes  (incluye el nombre del archivo)
      - 'comment'    → comment.content          (incluye el nombre del archivo)
    """
    from datetime import datetime, timezone
    from itcj2.apps.helpdesk.models.ticket import Ticket
    from itcj2.apps.helpdesk.models.comment import Comment

    try:
        att_type = attachment.attachment_type

        if att_type == "ticket":
            ticket = db.get(Ticket, attachment.ticket_id)
            if ticket is not None:
                ticket.description = (ticket.description or "") + (
                    f"\n\nSe adjuntó una imagen. (eliminado el {deletion_date})"
                )

        elif att_type == "resolution":
            ticket = db.get(Ticket, attachment.ticket_id)
            if ticket is not None:
                ticket.resolution_notes = (ticket.resolution_notes or "") + (
                    f"\n\nSe adjuntó archivo: {attachment.original_filename}"
                    f" (eliminado el {deletion_date})"
                )

        elif att_type == "comment" and attachment.comment_id:
            comment = db.get(Comment, attachment.comment_id)
            if comment is not None:
                comment.content = (comment.content or "") + (
                    f"\n\nSe adjuntó archivo: {attachment.original_filename}"
                    f" (eliminado el {deletion_date})"
                )
                comment.updated_at = datetime.now(timezone.utc)

    except Exception as e:
        logger.warning(
            f"[cleanup] No se pudo registrar nota de auditoría para "
            f"attachment {attachment.id}: {e}"
        )


# ---------------------------------------------------------------------------
# convert_document
# ---------------------------------------------------------------------------

@celery_app.task(
    bind=True,
    base=LoggedTask,
    name="itcj2.tasks.helpdesk_tasks.convert_document",
    max_retries=1,
    default_retry_delay=30,
    soft_time_limit=60,
    time_limit=90,
)
def convert_document(
    self,
    task_run_id: int,
    ticket_id: int,
    doc_type: str,
    notify_user_id: int,
) -> dict:
    """
    Genera la solicitud de mantenimiento o la orden de trabajo de un ticket
    como PDF en segundo plano usando LibreOffice.

    Al terminar, guarda el archivo en instance/apps/helpdesk/exports/ y
    crea una notificación con el enlace de descarga para el usuario.

    Args:
        task_run_id:     ID del TaskRun creado por la API.
        ticket_id:       ID del Ticket para el cual generar el documento.
        doc_type:        "solicitud" | "orden_trabajo"
        notify_user_id:  ID del usuario que recibirá la notificación.

    Returns:
        dict con claves: pdf_path, pages, doc_type, ticket_number, download_url
    """
    from datetime import datetime as _dt
    from itcj2.database import SessionLocal
    from itcj2.apps.helpdesk.models.ticket import Ticket
    from itcj2.apps.helpdesk.services.document_service import (
        generate_solicitud_pdf,
        generate_orden_trabajo_pdf,
    )

    if doc_type not in ("solicitud", "orden_trabajo"):
        raise ValueError(f"doc_type inválido: '{doc_type}'. Use 'solicitud' u 'orden_trabajo'.")

    logger.info(
        f"[convert_document] Iniciando — ticket_id={ticket_id}, "
        f"doc_type={doc_type}, task_run_id={task_run_id}"
    )

    self.update_progress(task_run_id, current=0, total=3, message="Cargando ticket...")

    # ── Paso 1: cargar ticket con relaciones dentro de la sesión ─────────
    with SessionLocal() as db:
        ticket = db.get(Ticket, ticket_id)
        if ticket is None:
            raise ValueError(f"Ticket {ticket_id} no encontrado")

        self.update_progress(
            task_run_id, current=1, total=3, message="Generando documento PDF..."
        )

        # Generamos el PDF dentro del contexto de la sesión para que las
        # relaciones (lazy-load) estén disponibles sin DetachedInstanceError.
        if doc_type == "solicitud":
            pdf_buffer = generate_solicitud_pdf(ticket)
            display_type = "Solicitud de Mantenimiento"
        else:
            pdf_buffer = generate_orden_trabajo_pdf(ticket)
            display_type = "Orden de Trabajo"

        ticket_number = ticket.ticket_number  # Capturar antes de cerrar sesión

    # ── Paso 2: guardar PDF en disco ─────────────────────────────────────
    self.update_progress(task_run_id, current=2, total=3, message="Guardando archivo...")

    timestamp = _dt.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"{doc_type}_{ticket_number}_{timestamp}.pdf"
    exports_dir = _get_exports_dir()
    pdf_path = os.path.join(exports_dir, filename)

    with open(pdf_path, "wb") as f:
        f.write(pdf_buffer.read())

    # ── Paso 3: contar páginas (opcional, requiere pypdf) ─────────────────
    pages = _count_pdf_pages(pdf_path)

    download_url = f"/api/helpdesk/exports/{filename}"

    # ── Paso 4: notificar al usuario ─────────────────────────────────────
    _push_user_notification(
        user_id=notify_user_id,
        title=f"{display_type} lista",
        body=f"El PDF del ticket {ticket_number} está disponible para descargar.",
        link=download_url,
    )

    result = {
        "pdf_path": pdf_path,
        "pages": pages,
        "doc_type": doc_type,
        "ticket_number": ticket_number,
        "download_url": download_url,
    }
    logger.info(f"[convert_document] Completado: {result}")
    return result


def _count_pdf_pages(pdf_path: str) -> int | None:
    """Cuenta páginas de un PDF usando pypdf. Devuelve None si no está disponible."""
    try:
        from pypdf import PdfReader
        reader = PdfReader(pdf_path)
        return len(reader.pages)
    except Exception:
        return None


# ---------------------------------------------------------------------------
# export_inventory_report
# ---------------------------------------------------------------------------

@celery_app.task(
    bind=True,
    base=LoggedTask,
    name="itcj2.tasks.helpdesk_tasks.export_inventory_report",
    max_retries=1,
    default_retry_delay=60,
    soft_time_limit=300,
    time_limit=360,
    queue="reports",
)
def export_inventory_report(
    self,
    task_run_id: int,
    filters: dict,
    format: str,
    requested_by_user_id: int,
) -> dict:
    """
    Genera un reporte de equipos de inventario en CSV o XLSX en segundo plano.

    Guarda el archivo en instance/apps/helpdesk/exports/ y crea una notificación
    con el enlace de descarga para el usuario solicitante.

    Args:
        task_run_id:          ID del TaskRun creado por la API.
        filters:              Filtros del reporte (ver InventoryReportsService).
        format:               "csv" | "xlsx"
        requested_by_user_id: ID del usuario que recibirá la notificación.

    Returns:
        dict con claves: file_path, rows, format, download_url
    """
    from datetime import datetime as _dt
    from itcj2.database import SessionLocal
    from itcj2.apps.helpdesk.services.inventory_reports_service import InventoryReportsService

    if format not in ("csv", "xlsx"):
        raise ValueError(f"format inválido: '{format}'. Use 'csv' o 'xlsx'.")

    logger.info(
        f"[export_inventory_report] Iniciando — format={format}, "
        f"filters={filters}, task_run_id={task_run_id}"
    )

    self.update_progress(
        task_run_id, current=0, total=3, message="Consultando datos de inventario..."
    )

    # ── Paso 1: obtener todos los items (sin paginación) ─────────────────
    with SessionLocal() as db:
        report_data = InventoryReportsService.get_equipment_report(
            db, {**filters, "page": 1, "per_page": 50_000}
        )

    items = report_data["items"]
    row_count = len(items)
    logger.info(f"[export_inventory_report] {row_count} equipos obtenidos")

    self.update_progress(
        task_run_id, current=1, total=3,
        message=f"Generando archivo {format.upper()} con {row_count} equipos...",
    )

    # ── Paso 2: generar el archivo ────────────────────────────────────────
    timestamp = _dt.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"inventario_{timestamp}.{format}"
    exports_dir = _get_exports_dir()
    file_path = os.path.join(exports_dir, filename)

    if format == "csv":
        _write_equipment_csv(file_path, items)
    else:
        _write_equipment_xlsx(file_path, items)

    download_url = f"/api/helpdesk/exports/{filename}"

    self.update_progress(task_run_id, current=2, total=3, message="Notificando al usuario...")

    # ── Paso 3: notificar al usuario ─────────────────────────────────────
    _push_user_notification(
        user_id=requested_by_user_id,
        title="Reporte de inventario listo",
        body=f"El reporte con {row_count} equipos está disponible para descargar ({format.upper()}).",
        link=download_url,
    )

    result = {
        "file_path": file_path,
        "rows": row_count,
        "format": format,
        "download_url": download_url,
    }
    logger.info(f"[export_inventory_report] Completado: {result}")
    return result


# ── Helpers de escritura de archivo ─────────────────────────────────────────

_CSV_HEADERS = [
    "No. Inventario", "Categoría", "Marca", "Modelo", "No. Serie",
    "Departamento", "Asignado a", "Ubicación", "Estado",
    "Fecha Adquisición", "Vencimiento Garantía",
    "Último Mantenimiento", "Próx. Mantenimiento", "Notas",
]


def _item_to_row(item: dict) -> list:
    dept = item.get("department") or {}
    user = item.get("assigned_to_user") or {}
    cat = item.get("category") or {}
    return [
        item.get("inventory_number", ""),
        cat.get("name", ""),
        item.get("brand", ""),
        item.get("model", ""),
        item.get("supplier_serial", ""),
        dept.get("name", ""),
        user.get("full_name", "Sin asignar"),
        item.get("location_detail", ""),
        item.get("status", ""),
        item.get("acquisition_date", ""),
        item.get("warranty_expiration", ""),
        item.get("last_maintenance_date", ""),
        item.get("next_maintenance_date", ""),
        item.get("notes", ""),
    ]


def _write_equipment_csv(file_path: str, items: list[dict]) -> None:
    import csv
    with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(_CSV_HEADERS)
        for item in items:
            writer.writerow(_item_to_row(item))


def _write_equipment_xlsx(file_path: str, items: list[dict]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # Cabecera con estilo
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F497D")

    for col_idx, header in enumerate(_CSV_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Filas de datos
    for row_idx, item in enumerate(items, start=2):
        for col_idx, value in enumerate(_item_to_row(item), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Ajuste automático de ancho de columna
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    wb.save(file_path)
