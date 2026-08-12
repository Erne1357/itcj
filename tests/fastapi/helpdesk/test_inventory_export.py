"""Exportación de inventario a Excel — `InventoryExportService`
(itcj2/apps/helpdesk/services/inventory_export_service.py) y el endpoint
GET /api/help-desk/v2/inventory/items/export.xlsx que lo expone.

El servicio se prueba DIRECTO (sin HTTP) para poder abrir el workbook con
openpyxl y verificar hojas/encabezados/auto_filter/freeze_panes/formato
condicional sin pasar por la capa HTTP. Los tests de content-type/
Content-Disposition sí van por HTTP (es lo único que el service no controla).
"""
from datetime import date, datetime, timedelta
from io import BytesIO
from uuid import uuid4

import time

import jwt
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

import itcj2.models  # noqa: F401 (resuelve mappers)
from itcj2.apps.helpdesk.models import InventoryItem
from itcj2.apps.helpdesk.services.inventory_export_service import InventoryExportService
from itcj2.config import get_settings
from itcj2.core.models.app import App
from itcj2.core.models.department import Department
from itcj2.core.models.permission import Permission
from itcj2.core.models.position import Position, PositionAppPerm, UserPosition
from itcj2.core.models.user import User
from itcj2.database import get_db
from itcj2.main import create_app

from ._catalog import ensure_inventory_category

ITEMS_SUBTREE = "helpdesk.inventory.api.read.subtree"


# ─────────────────────────── infraestructura de test ───────────────────────────

@pytest.fixture()
def client(db_session):
    app = create_app()
    app.dependency_overrides[get_db] = lambda: db_session
    with TestClient(app) as c:
        yield c
    app.dependency_overrides.clear()


def _jwt_cookie(user_id: int) -> dict:
    settings = get_settings()
    now = int(time.time())
    payload = {
        "sub": str(user_id), "role": None, "cn": None, "name": "Test",
        "iat": now, "exp": now + 24 * 3600,
    }
    token = jwt.encode(payload, settings.SECRET_KEY, algorithm="HS256")
    return {"Cookie": f"itcj_token={token}"}


def _helpdesk(db) -> App:
    app = db.query(App).filter_by(key="helpdesk").first()
    assert app is not None, "helpdesk app debe existir en la BD dev"
    return app


def _perm(db, app, code) -> Permission:
    p = db.query(Permission).filter_by(app_id=app.id, code=code).first()
    if not p:
        p = Permission(app_id=app.id, code=code, name=code)
        db.add(p)
        db.commit()
        db.refresh(p)
    return p


def _dept(db, code, parent_id=None) -> Department:
    d = Department(code=code, name=code, parent_id=parent_id, is_active=True)
    db.add(d)
    db.commit()
    db.refresh(d)
    return d


def _user(db, last) -> User:
    u = User(first_name="T", last_name=last, is_active=True)
    db.add(u)
    db.commit()
    db.refresh(u)
    return u


def _item(db, number, department, registered_by, **kwargs) -> InventoryItem:
    status = kwargs.pop("status", "ACTIVE")
    it = InventoryItem(
        inventory_number=number,
        category_id=ensure_inventory_category(db).id,
        department_id=department.id,
        registered_by_id=registered_by.id,
        status=status,
        is_active=True,
        **kwargs,
    )
    db.add(it)
    db.commit()
    db.refresh(it)
    return it


def _grant(db, user, department, *, code) -> None:
    codes = [code] if isinstance(code, str) else list(code)
    app = _helpdesk(db)
    pos = Position(code=f"iexp_{user.id}_{uuid4().hex[:8]}", title="Jefe", department_id=department.id,
                    is_active=True, allows_multiple=True)
    db.add(pos)
    db.commit()
    db.refresh(pos)
    db.add(UserPosition(user_id=user.id, position_id=pos.id,
                         start_date=date.today() - timedelta(days=1), is_active=True))
    for c in codes:
        perm = _perm(db, app, c)
        db.add(PositionAppPerm(position_id=pos.id, app_id=app.id, perm_id=perm.id, allow=True))
    db.commit()


def _tree(db, prefix):
    """root -> mine -> child ; root -> sibling (rama hermana)."""
    root = _dept(db, f"{prefix}_root")
    mine = _dept(db, f"{prefix}_mine", root.id)
    child = _dept(db, f"{prefix}_child", mine.id)
    sibling = _dept(db, f"{prefix}_sibling", root.id)
    return root, mine, child, sibling


def _full_access(db, prefix):
    """Departamento AISLADO (recién creado, sin equipos preexistentes) + usuario
    con `.read.subtree` anclado ahí. Deliberadamente NO usa `.read.all`: contra
    la Postgres real de dev (harness de `tests/fastapi/conftest.py`) un acceso
    verdaderamente global vería TODO el inventario ya sembrado (miles de filas
    reales) a través de la MISMA sesión del test, contaminando cualquier assert
    de contenido/estructura del workbook. Anclar en un depto nuevo consigue el
    mismo "ve todo lo que yo cree" sin esa fuga.
    Devuelve (department, user, cookie)."""
    dept = _dept(db, f"{prefix}_dept")
    user = _user(db, f"{prefix}Full")
    _grant(db, user, dept, code=ITEMS_SUBTREE)
    return dept, user, _jwt_cookie(user.id)


def _user_dict(user) -> dict:
    """Payload mínimo compatible con `visible_department_ids`/`is_global_admin`
    (mismo shape que deja el JWT en `request.state.current_user`)."""
    return {"sub": str(user.id), "role": None}


# ─────────────────────────────── estructura del workbook ───────────────────────────────

class TestExportServiceStructure:
    def test_two_sheets_no_metadata_sheet(self, db_session):
        dept, user, _ = _full_access(db_session, "exs1")
        _item(db_session, "EXS1-A", dept, user)

        buf, filename = InventoryExportService.export_items(db_session, _user_dict(user), {})

        wb = load_workbook(buf)
        assert wb.sheetnames == ["Inventario", "Resumen"]
        assert filename.startswith("inventario_")
        assert filename.endswith(".xlsx")

    def test_headers_row(self, db_session):
        dept, user, _ = _full_access(db_session, "exs2")
        _item(db_session, "EXS2-A", dept, user)

        buf, _ = InventoryExportService.export_items(db_session, _user_dict(user), {})

        ws = load_workbook(buf)["Inventario"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert headers[0] == "N° Inventario"
        assert "Días de garantía" in headers
        assert "Días sin verificar" in headers
        assert "Registrado el" in headers

    def test_auto_filter_and_freeze_panes(self, db_session):
        dept, user, _ = _full_access(db_session, "exs3")
        _item(db_session, "EXS3-A", dept, user)
        _item(db_session, "EXS3-B", dept, user)

        buf, _ = InventoryExportService.export_items(db_session, _user_dict(user), {})

        ws = load_workbook(buf)["Inventario"]
        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref == "A1:U3"

    def test_conditional_formatting_present_for_normal_size(self, db_session):
        dept, user, _ = _full_access(db_session, "exs4")
        _item(db_session, "EXS4-A", dept, user, status="DAMAGED")

        buf, _ = InventoryExportService.export_items(db_session, _user_dict(user), {})

        ws = load_workbook(buf)["Inventario"]
        assert len(list(ws.conditional_formatting)) > 0

    def test_summary_has_total_and_active_filters_text(self, db_session):
        dept, user, _ = _full_access(db_session, "exs5")
        _item(db_session, "EXS5-A", dept, user, brand="Dell Inspiron")

        buf, _ = InventoryExportService.export_items(db_session, _user_dict(user), {"brand": "dell"})

        ws = load_workbook(buf)["Resumen"]
        blob = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
        assert "Total exportado: 1" in blob
        assert "Marca: dell" in blob

    def test_summary_has_a_bar_chart_per_block(self, db_session):
        dept, user, _ = _full_access(db_session, "exs6")
        _item(db_session, "EXS6-A", dept, user)
        _item(db_session, "EXS6-B", dept, user, status="MAINTENANCE")

        buf, _ = InventoryExportService.export_items(db_session, _user_dict(user), {})

        ws = load_workbook(buf)["Resumen"]
        # Un bloque por: estado, categoría, departamento, asignación, garantía, verificación.
        assert len(ws._charts) >= 6


# ─────────────────────────────── filtros respetados ───────────────────────────────

class TestExportServiceFilters:
    def test_respects_query_filters(self, db_session):
        dept, user, _ = _full_access(db_session, "exf1")
        _item(db_session, "EXF1-KEEP", dept, user, brand="Dell Inspiron")
        _item(db_session, "EXF1-DROP", dept, user, brand="HP LaserJet")

        buf, _ = InventoryExportService.export_items(db_session, _user_dict(user), {"brand": "dell"})

        ws = load_workbook(buf)["Inventario"]
        numbers = [row[0].value for row in ws.iter_rows(min_row=2)]
        assert numbers == ["EXF1-KEEP"]

    def test_respects_status_filter(self, db_session):
        dept, user, _ = _full_access(db_session, "exf2")
        _item(db_session, "EXF2-ACTIVE", dept, user, status="ACTIVE")
        _item(db_session, "EXF2-DAMAGED", dept, user, status="DAMAGED")

        buf, _ = InventoryExportService.export_items(db_session, _user_dict(user), {"status": "DAMAGED"})

        ws = load_workbook(buf)["Inventario"]
        numbers = [row[0].value for row in ws.iter_rows(min_row=2)]
        assert numbers == ["EXF2-DAMAGED"]


# ─────────────────────────────── scope acotado ───────────────────────────────

class TestExportServiceScope:
    def test_subtree_scope_excludes_sibling_branch(self, db_session):
        root, mine, child, sibling = _tree(db_session, "exsc1")
        boss = _user(db_session, "ExscBoss1")
        registrar = _user(db_session, "ExscReg1")
        _grant(db_session, boss, mine, code=ITEMS_SUBTREE)
        _item(db_session, "EXSC1-MINE", mine, registrar)
        _item(db_session, "EXSC1-SIB", sibling, registrar)

        buf, _ = InventoryExportService.export_items(db_session, {"sub": str(boss.id), "role": None}, {})

        ws = load_workbook(buf)["Inventario"]
        numbers = [row[0].value for row in ws.iter_rows(min_row=2)]
        assert numbers == ["EXSC1-MINE"]

    def test_department_param_outside_scope_exports_empty(self, db_session):
        root, mine, child, sibling = _tree(db_session, "exsc2")
        boss = _user(db_session, "ExscBoss2")
        registrar = _user(db_session, "ExscReg2")
        _grant(db_session, boss, mine, code=ITEMS_SUBTREE)
        _item(db_session, "EXSC2-SIB", sibling, registrar)

        buf, _ = InventoryExportService.export_items(
            db_session, {"sub": str(boss.id), "role": None}, {"department": str(sibling.id)},
        )

        ws = load_workbook(buf)["Inventario"]
        assert list(ws.iter_rows(min_row=2)) == []


# ─────────────────────────────── degradación (>20k filas) ───────────────────────────────

class TestExportDegradation:
    def test_large_result_skips_conditional_formatting_and_notes_it(self, db_session, monkeypatch):
        monkeypatch.setattr(
            "itcj2.apps.helpdesk.services.inventory_export_service._MAX_CONDITIONAL_ROWS", 1,
        )
        dept, user, _ = _full_access(db_session, "exd1")
        _item(db_session, "EXD1-A", dept, user)
        _item(db_session, "EXD1-B", dept, user)

        buf, _ = InventoryExportService.export_items(db_session, _user_dict(user), {})

        wb = load_workbook(buf)
        ws = wb["Inventario"]
        assert len(list(ws.conditional_formatting)) == 0

        summary_blob = " ".join(
            str(c.value) for row in wb["Resumen"].iter_rows() for c in row if c.value is not None
        )
        assert "formato condicional" in summary_blob.lower()


# ─────────────────────────────── HTTP: content-type / Content-Disposition ───────────────────────────────

class TestExportHttp:
    def test_content_type_and_disposition(self, client, db_session):
        dept, user, cookie = _full_access(db_session, "exh1")
        _item(db_session, "EXH1-A", dept, user)

        resp = client.get("/api/help-desk/v2/inventory/items/export.xlsx", headers=cookie)

        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        disposition = resp.headers.get("content-disposition", "")
        assert "attachment" in disposition
        assert ".xlsx" in disposition

        wb = load_workbook(BytesIO(resp.content))
        assert wb.sheetnames == ["Inventario", "Resumen"]

    def test_requires_helpdesk_app_access_same_as_list(self, client, db_session):
        stranger = _user(db_session, "ExhStranger")

        resp = client.get("/api/help-desk/v2/inventory/items/export.xlsx", headers=_jwt_cookie(stranger.id))

        assert resp.status_code == 403

    def test_accepts_same_param_names_as_the_page_form(self, client, db_session):
        dept, user, cookie = _full_access(db_session, "exh2")
        _item(db_session, "EXH2-KEEP", dept, user, brand="Dell")
        _item(db_session, "EXH2-DROP", dept, user, brand="HP")

        resp = client.get("/api/help-desk/v2/inventory/items/export.xlsx?brand=dell", headers=cookie)

        assert resp.status_code == 200
        ws = load_workbook(BytesIO(resp.content))["Inventario"]
        numbers = [row[0].value for row in ws.iter_rows(min_row=2)]
        assert numbers == ["EXH2-KEEP"]
